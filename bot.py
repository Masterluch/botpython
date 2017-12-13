import telebot
import openpyxl
import settings

work_sheets = openpyxl.load_workbook(settings.file)

# work_sheet_9a = openpyxl.load_workbook("C:\\Users\\Андрей\\Desktop\\Расписание.xlsx")['9"А"']
# work_sheet_9b = openpyxl.load_workbook("C:\\Users\\Андрей\\Desktop\\Расписание.xlsx")['9"Б"']
# work_sheet_9v = openpyxl.load_workbook("C:\\Users\\Андрей\\Desktop\\Расписание.xlsx")['9"В"']
# work_sheet_8a = openpyxl.load_workbook("C:\\Users\\Андрей\\Desktop\\Расписание.xlsx")['8"А"']

statistics = {}

table = {
    "ПОНЕДЕЛЬНИК": 'B',
    "ВТОРНИК": 'C',
    "СРЕДА": 'D',
    "ЧЕТВЕРГ": 'E',
    "ПЯТНИЦА": 'F',
}

def make_requets_to_table(dayWeek, lessonNumber):
    res = ''
    res += table[dayWeek]
    res += str(int(lessonNumber)+1)
    return res

bot = telebot.TeleBot(settings.apikey)

@bot.message_handler(func=lambda m: m.from_user.username == 'Masterluch' and m.text == '/statistics')
def admins_statistics_get(msg):
    print('admin would statistics')
    for k, v in statistics.items():
        bot.reply_to(msg, k + ": " + str(v))


@bot.message_handler(regexp="расписание \d\w на \w+ \d урок")
def table_class_letter_day_lesson(msg):
    print("HERE")
    tokenized = msg.text.upper().split(" ")
    print(tokenized)
    day = ""
    lesson = ""
    try:
        print("someone want table for " + tokenized[1])
        temp_work_sheet = work_sheets[tokenized[1]]
        lesson = temp_work_sheet[make_requets_to_table(tokenized[3], tokenized[4])].value
        try:
            statistics[tokenized[1]] += 1
        except KeyError:
            statistics[tokenized[1]] = 1
    except KeyError:
        print("key error")
        return
    except IndexError:
        print("index error")
        return
    if lesson == None:
        print("no lesson")
        return
    bot.reply_to(msg, lesson)

bot.polling()